import openpyxl
import win32com.client

module_name = '模块1'
method_name = '按钮1_Click'

years = []
cals = [[1], [2], [3], [4], [1, 2], [1, 3], [1, 4], [2, 3], [2, 4], [3, 4],
        [1, 2, 3], [1, 2, 4], [1, 3, 4], [2, 3, 4], [1, 2, 3, 4]]

needRows = []


def handle_vba(filename):
    xls = win32com.client.Dispatch("Excel.Application")
    xls.Application.visible = False

    # print('转换开始')
    ####调用vba程序。需要安装win32com库

    try:
        book = xls.workbooks.Open(filename)  ##存储vba代码的文件
        sheet = book.Worksheets('sbm效率计算')
        # xl.Application.Run("excelsheet.xlsm!modulename.macroname")
        url = module_name + "." + method_name
        sheet.Application.Run(url)
        ##开始调用vba宏
        # status = sheet.Application.ExecuteExcel4Macro('sbm结果计算宏文件1.xlsm!按钮1_Click()')
        # print(status)
        book.Close(SaveChanges=True)
    except Exception as ex:
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print(message)

    # book.Close(SaveChanges=True)
    xls.Application.Quit()


def copy_value_to_cal_sheet(fromSheet, from_index, toSheet):
    for index in range(71):
        cal_row = index + 4
        if index < len(from_index):
            for column in range(8):
                tmp = fromSheet.cell(row=from_index[index], column=column + 2).value
                toSheet.cell(row=cal_row, column=column + 1).value = tmp
        else:
            toSheet.cell(row=cal_row, column=4).value = 1000
            toSheet.cell(row=cal_row, column=5).value = 1000
            toSheet.cell(row=cal_row, column=6).value = 1000
            toSheet.cell(row=cal_row, column=7).value = 1000
            toSheet.cell(row=cal_row, column=8).value = 0.001
            toSheet.cell(row=cal_row, column=9).value = 1000


def handler_var(switch_year, level, filename, initMoveRow):
    workbook = openpyxl.load_workbook(filename=filename, keep_vba=True)
    ori_sheet = workbook["汇总"]

    # 条件判断
    needRows = []
    for row in range(3, 711):
        city = ori_sheet.cell(row=row, column=1).value
        year = ori_sheet.cell(row=row, column=3).value
        if city in level and year == switch_year:
            needRows.append(row)
    # 拷贝数据
    copy_value_to_cal_sheet(ori_sheet, needRows, workbook["sbm效率规模报酬可变"])
    # copy_value_to_cal_sheet("sbm效率规模报酬可变")
    # copy_value_to_cal_sheet("sbm效率计算")
    workbook.active = workbook['sbm效率规模报酬可变']
    workbook.save(filename=filename)
    workbook.close()
    # 计算数据
    handle_vba(filename)
    # 读取结果
    workbook = openpyxl.load_workbook(filename=filename, keep_vba=True, data_only=True)
    result_sheet_b = workbook["sbm效率规模报酬可变"]
    # # 条件判断
    tmpResult = []
    for row_index in range(len(needRows)):
        tmp = []
        for col_index in range(14):
            row = row_index + 4
            column = col_index + 22
            tmp.append(result_sheet_b.cell(row=row, column=column).value)
        tmpResult.append(tmp)

    workbook.close()
    # print(tmpResult)
    workbook = openpyxl.load_workbook(filename=filename, keep_vba=True)
    to_sheet_b1 = workbook["规模报酬可变"]
    to_sheet_b2 = workbook["非期望效率汇总"]
    for row_index, row in enumerate(tmpResult):
        r_row = row_index + initMoveRow
        to_sheet_b2.cell(row=r_row, column=5).value = row[0]
        for col_index, cell in enumerate(row):
            to_sheet_b1.cell(row=r_row, column=col_index + 4).value = cell
    workbook.save(filename=filename)
    workbook.close()


def handler_static(switch_year, level, filename, initMoveRow):
    print("start static")
    workbook = openpyxl.load_workbook(filename=filename, keep_vba=True)
    ori_sheet = workbook["汇总"]

    # 条件判断
    needRows = []
    for row in range(3, 711):
        city = ori_sheet.cell(row=row, column=1).value
        year = ori_sheet.cell(row=row, column=3).value
        if city in level and year == switch_year:
            needRows.append(row)
    # 拷贝数据
    copy_value_to_cal_sheet(ori_sheet, needRows, workbook["sbm效率计算"])
    # copy_value_to_cal_sheet("sbm效率规模报酬可变")
    # copy_value_to_cal_sheet("sbm效率计算")
    workbook.active = workbook['sbm效率计算']
    workbook.save(filename=filename)
    workbook.close()
    # 计算数据
    handle_vba(filename)
    # 读取结果
    workbook = openpyxl.load_workbook(filename=filename, keep_vba=True, data_only=True)
    result_sheet_b = workbook["sbm效率计算"]
    # # 条件判断
    tmpResult = []
    for row_index in range(len(needRows)):
        tmp = []
        for col_index in range(14):
            row = row_index + 4
            column = col_index + 22
            tmp.append(result_sheet_b.cell(row=row, column=column).value)
        tmpResult.append(tmp)

    workbook.close()
    # print(tmpResult)
    workbook = openpyxl.load_workbook(filename=filename, keep_vba=True)
    to_sheet_b1 = workbook["规模报酬不变"]
    to_sheet_b2 = workbook["非期望效率汇总"]
    for row_index, row in enumerate(tmpResult):
        row_1 = row_index + initMoveRow
        to_sheet_b2.cell(row=row_1, column=4).value = row[0]
        for col_index, cell in enumerate(row):
            to_sheet_b1.cell(row=row_1, column=col_index + 4).value = cell
    workbook.save(filename=filename)
    workbook.close()
    return needRows


def handler_result_title(d_year, n_rows, filename, initMoveRow):
    print("start result")
    workbook = openpyxl.load_workbook(filename=filename, keep_vba=True)
    ori_sheet = workbook["汇总"]
    a_sheet = workbook["规模报酬可变"]
    b_sheet = workbook["规模报酬不变"]
    c_sheet = workbook["非期望效率汇总"]
    for index, row in enumerate(n_rows):
        col_1_Value = index + 1
        row_A = initMoveRow + index
        a_sheet.cell(row=row_A, column=1).value = col_1_Value
        b_sheet.cell(row=row_A, column=1).value = col_1_Value
        c_sheet.cell(row=row_A, column=1).value = col_1_Value
        a_sheet.cell(row=row_A, column=2).value = d_year
        b_sheet.cell(row=row_A, column=2).value = d_year
        c_sheet.cell(row=row_A, column=2).value = d_year
        city_name = ori_sheet.cell(row=row, column=4).value
        a_sheet.cell(row=row_A, column=3).value = city_name
        b_sheet.cell(row=row_A, column=3).value = city_name
        c_sheet.cell(row=row_A, column=3).value = city_name
    workbook.save(filename=fileName)
    workbook.close()


def handler_del_row(level, filename):
    print("start delete row")
    workbook = openpyxl.load_workbook(filename=filename, keep_vba=True)
    ori_sheet = workbook["汇总"]
    needRows = []
    for row in range(3, 711):
        city = ori_sheet.cell(row=row, column=1).value
        if city in level:
            needRows.append(row)
    a_sheet = workbook["规模报酬可变"]
    b_sheet = workbook["规模报酬不变"]
    c_sheet = workbook["非期望效率汇总"]
    for index in range(713, len(needRows) + 3, -1):
        a_sheet.delete_rows(index)
        b_sheet.delete_rows(index)
        c_sheet.delete_rows(index)
    workbook.save(filename=filename)
    workbook.close()


if __name__ == '__main__':
    for level_index, level in enumerate(cals):
        initMoveRow = 4
        seq = level_index + 1
        fileName = r"D:\a=tmm\cai\sbm结果计算宏文件 - 副本 (" + str(seq) + ").xlsm"
        for i in range(10):
            y = 2010 + i
            print("start" + str(fileName) + ":" + str(y))
            handler_var(y, level, fileName, initMoveRow)
            n_rows = handler_static(y, level, fileName, initMoveRow)
            handler_result_title(y, n_rows, fileName, initMoveRow)
            initMoveRow += len(n_rows)
        handler_del_row(level, fileName)

    # copy_value_to_cal_sheet("sbm效率计算")
    # workbook.active = workbook['sbm效率计算']
    # workbook.save(filename=fileName)
    # handle_vba()
    #
# init_year(2010, 10)
# print(years)
#
