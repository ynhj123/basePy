import numbers

import openpyxl

if __name__ == '__main__':
    # 打开Excel文件
    workbook = openpyxl.load_workbook('F:\\Users\\副本碳排放(最终版）(1).xlsx')

    # 获取sheet对象
    worksheet = workbook["数值新"]

    # 遍历sheet中的每个单元格
    for row in range(5, worksheet.max_row):
        for col in range(4, worksheet.max_column):
            value = worksheet.cell(row=row, column=col).value
            value += 1
            if not isinstance(value, numbers.Number):
                print(str(row) + ":" + str(col) + "=" + str(value) + str(isinstance(value, numbers.Number)))
